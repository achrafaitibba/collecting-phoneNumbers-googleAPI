import googlemaps
from openpyxl import Workbook, load_workbook
from datetime import datetime


def collectPhoneNumbers(keyword, city):
    # Initialize Google Maps client with API key
    gmaps = googlemaps.Client(key='')

    # Search for places with keyword and city
    places_result = gmaps.places(query=keyword, location=city, radius=5000)

    # Create a new or load existing workbook
    try:
        excel_file = load_workbook(filename="phone_numbers1.xlsx")
        sheet = excel_file.active
    except:
        excel_file = Workbook()
        sheet = excel_file.active
        sheet.append(['Name', 'Address', 'City', 'Phone', 'Website', 'Date'])

    # Get list of existing phone numbers
    existing_phones = [row[3].value for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4)]

    # Loop through each place and get details
    for place in places_result['results']:
        # Check if phone number already exists
        if place.get('formatted_phone_number', '') in existing_phones:
            continue

        # Get place details using place_id
        place_id = place['place_id']
        details = gmaps.place(place_id=place_id)['result']

        # Extract required details
        name = details.get('name', '')
        address = details.get('formatted_address', '')
        city = details.get('formatted_address', '').split(',')[-1].strip()
        phone = details.get('formatted_phone_number', '')
        website = details.get('website', '')
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Append details to the sheet
        sheet.append([name, address, city, phone, website, date])

    # Save the updated workbook
    excel_file.save("phone_numbers1.xlsx")


# Example usage
collectPhoneNumbers("Boulangerie", "Rabat")
